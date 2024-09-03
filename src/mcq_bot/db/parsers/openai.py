import asyncio
import json
import logging
from pathlib import Path

from openai import AsyncOpenAI
from openpyxl import load_workbook
from openpyxl.cell import Cell
from tqdm import tqdm

from mcq_bot.db.db_types import ProcessedRow
from mcq_bot.db.parsers.base import BaseParser
from mcq_bot.settings import Settings

_logger = logging.getLogger(__name__)


class LlmFailedToParseError(Exception):
    """LLM failed to parse the question."""


class OpenAiParser(BaseParser):
    """
    Parses an excel file containing questions, answers and explanations in each row.

    Each row is passed as a JSON dict to the LLM, with the keys being the cells in the first row of the file (the headers).

    As a result, the format can be variable.

    Note: Use a new instance for each file.
    """

    MODEL = "gpt-4o-mini"
    USER_MESSAGE_PREFIX = "Parse this dictionary containing questions, answer options, answers and explanations into the provided schema. Don't repeat the answer choices in the question text. If the question is empty, you should refuse the request."

    def __init__(self, concurrent_requests: int = 10) -> None:
        """
        Return an OpenAIParser.

        concurrent_requests: Max simultaneous/pending requests to OpenAI.
        """
        self.client = AsyncOpenAI(api_key=Settings.OPENAI_API_KEY.get_secret_value())
        self.queue: asyncio.Queue[dict[str, str]] = asyncio.Queue()
        self.concurrent_requests = concurrent_requests
        self.raw_results: list[ProcessedRow | None] = []

    def _format_row_to_dict(
        self, row: tuple[Cell, ...], headers: tuple[Cell, ...]
    ) -> dict[str, str]:
        """
        Add headers to the row and returns a dict with header columns as the keys and the row values as the value.

        If there are less headers than rows, uses "Unnamed column {index of row}" as the key.
        """
        result: dict[str, str] = {}
        for idx, r in enumerate(row):
            if len(headers) >= idx and (header_val := headers[idx].value):
                key = str(header_val)
            else:
                key = f"Unnamed column {idx}"
            result[key] = str(r.value) if r.value else "Empty"
        return result

    async def _llm_parse_row(
        self, formatted_row: dict[str, str]
    ) -> ProcessedRow | None:
        """Try to parse an Openpyxl row with the llm. Returns None if failed."""
        try:
            completion = await self.client.beta.chat.completions.parse(
                model=self.MODEL,
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {
                        "role": "user",
                        "content": f"{json.dumps(formatted_row, indent=2)}\n{self.USER_MESSAGE_PREFIX}",
                    },
                ],
                response_format=ProcessedRow,
            )
            message = completion.choices[0].message
            parsed = message.parsed

            if parsed and message.parsed:
                return message.parsed
            else:
                _logger.error(
                    "Failed to parse row:\n%s\n\n%s",
                    json.dumps(formatted_row, indent=2),
                    message.refusal,
                )
                return None

        except Exception:
            _logger.exception(
                "Failed to send row for processing:\n%s\n\n%s",
                json.dumps(formatted_row, indent=2),
            )
            return None

    async def _worker(self, pbar: tqdm):
        while True:
            formatted_row = await self.queue.get()
            result = await self._llm_parse_row(formatted_row)
            self.raw_results.append(result)
            pbar.update()
            self.queue.task_done()

    async def _parse(self, path: Path) -> list[ProcessedRow]:
        wb = load_workbook(path)
        sheet = wb[wb.sheetnames[0]]
        processed_rows: list[ProcessedRow] = []
        headers = list(sheet.rows)[0]

        total = len(list(sheet.rows)) - 1  # Don't count the header row

        with tqdm(total=total, desc=path.name) as pbar:
            workers = [
                asyncio.create_task(self._worker(pbar))
                for _ in range(self.concurrent_requests)
            ]

            for row in sheet.iter_rows(min_row=2):
                formatted_row = self._format_row_to_dict(row, headers)
                await self.queue.put(formatted_row)

            await self.queue.join()
            [w.cancel() for w in workers]

            processed_rows = [r for r in self.raw_results if r is not None]

        return processed_rows

    def parse(self, path: Path) -> list[ProcessedRow]:
        """Process the excel file at path. Note: due to the nature of LLMs, some rows may fail to be parsed. They will not appear in the output."""

        loop = asyncio.get_event_loop()
        result = loop.run_until_complete(self._parse(path))

        return result
