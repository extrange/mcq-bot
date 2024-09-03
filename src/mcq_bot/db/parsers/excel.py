from pathlib import Path
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell

from mcq_bot.db.db_types import (
    VALID_ANSWER_LETTERS,
    AnswerKeys,
    AnswerType,
    NoCorrectAnswerException,
    ProcessedRow,
    QuestionType,
)

from .base import BaseParser
from .utils import validate_only_one_correct_answer


class ExcelParser(BaseParser):
    """
    Parse an excel file in a strict format (see `process_row`).

    Column indices expected format:
    0: Qn number
    1: Question text
    2-6: Answers
    7: Answer
    8: Explanation
    """

    def _extract_question(self, row: Tuple[Cell, ...]):
        question = QuestionType(
            text=str(row[1].value).strip(),
            explanation=str(row[8].value).strip(),
        )
        return question

    def _process_row(
        self, row: Tuple[Cell, ...], answer_keys: list[AnswerKeys]
    ) -> ProcessedRow:
        """
        Process a single row and return the question and the answer.
        """
        question = self._extract_question(row)
        correct_letter = str(row[7].value).upper().strip()

        if not correct_letter or correct_letter not in VALID_ANSWER_LETTERS:
            raise NoCorrectAnswerException(
                f"Invalid or no correct answer provided: {correct_letter=}"
            )

        answers: list[AnswerType] = []

        for idx, cell in enumerate(row[2:7]):
            val = str(cell.value).strip() if cell.value else None

            # If the cell is blank there is no answer for that index
            if not val:
                continue

            formatted_text = val[:1].upper() + val[1:]

            answer = AnswerType(
                is_correct=answer_keys[idx] == correct_letter,
                key=idx,
                text=formatted_text,
            )
            answers.append(answer)

        # There must be exactly one correct answer
        try:
            validate_only_one_correct_answer(answers)
        except NoCorrectAnswerException as e:
            raise NoCorrectAnswerException(
                f"Error while processing 'question' {question.text[:50]}...",
            ) from e

        return ProcessedRow(question=question, answers=answers)

    def parse(self, path: Path) -> list[ProcessedRow]:
        """
        Given an excel file containing questions, return the questions, explanations and answers.
        """

        wb = load_workbook(path)
        sheet = wb[wb.sheetnames[0]]

        processed_rows: list[ProcessedRow] = []

        for idx, r in enumerate(sheet.iter_rows(min_row=2)):
            try:
                # Skip rows without questions
                if r[1].value is None:
                    continue
                processed_rows.append(self._process_row(r, VALID_ANSWER_LETTERS))
            except NoCorrectAnswerException as e:
                raise NoCorrectAnswerException(
                    f"Error while processing row {idx + 2} for file {path}"
                ) from e

        return processed_rows
